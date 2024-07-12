from robot.api.deco import keyword
from openpyxl import load_workbook
import random
import string
from datetime import datetime, timedelta
import email
import imaplib


import math
@keyword
def fetch_testdata_by_id(file_path, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict

        # If the target_id is not found, you can raise an exception or return a specific value
        raise ValueError(f"Target ID '{target_id}' not found in the Excel file.")
    except Exception as e:
        print(f"Error reading Excel file: {e}")


@keyword
def get_trade_name():
    prefix = "Qa-"
    random_suffix = ''.join(random.choices(string.ascii_lowercase, k=6))
    trade_name = prefix + random_suffix
    return trade_name


@keyword
def get_legal_name():
    suffix = "pvt"
    random_prefix = ''.join(random.choices(string.ascii_lowercase, k=6))
    legal_name = random_prefix.capitalize() + " " + suffix
    return legal_name


@keyword
def get_website_name():
    domain = ''.join(random.choices(string.ascii_lowercase, k=random.randint(5, 10)))
    return f"www.{domain}.com"


@keyword
def generate_fake_email():
    return f"{''.join(random.choices(string.ascii_lowercase, k=8))}@gmail.com"


@keyword
def generate_fake_pan():
    prefix = ''.join(random.choices(string.ascii_uppercase, k=5))
    suffix = ''.join(random.choices(string.digits, k=4))
    checksum = ''.join(random.choices(string.ascii_uppercase, k=1))
    return prefix + suffix + checksum


@keyword
def generate_cin():
    section1 = 'L'
    section2 = ''.join(random.choices(string.digits, k=5))
    section3 = ''.join(random.choices(string.ascii_uppercase, k=2))
    section4 = ''.join(random.choices(string.digits, k=4))
    section5 = ''.join(random.choices(string.ascii_uppercase, k=3))
    section6 = ''.join(random.choices(string.digits, k=6))
    return f"{section1}{section2}{section3}{section4}{section5}{section6}"


@keyword
def generate_ba_code():
    return ''.join(random.choices('0123456789', k=8))


@keyword
def generate_mobile_number():
    return ''.join(random.choices('123456789', k=10))

def search_and_fetch_email(server, port, email_address, password, subject):
    # Connect to the email server
    mail = imaplib.IMAP4_SSL(server, port)
    mail.login(email_address, password)
    mail.select('inbox')

    # Search for emails with the specified subject
    result, data = mail.search(None, f'(SUBJECT "{subject}")')
    email_ids = data[0].split()

    # Fetch the latest email and extract the OTP from its body
    if email_ids:
        latest_email_id = email_ids[-1]
        result, data = mail.fetch(latest_email_id, '(RFC822)')
        raw_email = data[0][1].decode('utf-8')  # Decode bytes to string

        email_message = email.message_from_string(raw_email)  # Use message_from_string to parse email content

        # Extract the OTP from the email body
        email_body = ""
        if email_message.is_multipart():
            for part in email_message.walk():
                if part.get_content_type() == "text/plain":
                    email_body += part.get_payload(decode=True).decode()
        else:
            email_body = email_message.get_payload(decode=True).decode()

        return email_body
    # If no email with the specified subject is found or no OTP is found in the email body
    return None


@keyword
def get_password_from_mail_body(paragraph,word='Enterprise -'):
    lines = paragraph.split('\n')
    for line in lines:
        # Check if the word is in the line
        if word in line:
            # Find the position of the word in the line
            start = line.find(word)
            # Return the string from the word to the end of the line
            return line[start + len(word):].strip()
    return None

@keyword
def get_dates():
    # Get today's date
    today = datetime.today()
    # Extract the day of the month for today
    day_today = today.day

    # Subtract two days
    two_days_ago = today - timedelta(days=2)
    # Extract the day for two days ago
    day_two_days_ago = two_days_ago.day

    return day_two_days_ago, day_today

# def generate_mobile_number():
#     # The first digit of a phone number should not be 0 or 1
#     first_digit = random.randint(6, 9)
#     # Generate the remaining 9 digits
#     remaining_digits = [random.randint(0, 9) for _ in range(9)]
#     # Combine all digits into a string
#     phone_number = ''.join(map(str, [first_digit] + remaining_digits))
#     return phone_number

@keyword
def generate_random_address():
    return f"Street {random.randint(1, 999)}, {random.choice(['laxmi road', 'Mg raod', 'SB road'])}"


@keyword
def random_pin():
    pin_codes = ['411001', '411002', '411021', '411023', '411044', '410506', '412115', '412210', '412206', '412305',
                 '412212', '410503', '413801', '413102', '413106', '410502', '410501']
    return random.choice(pin_codes)

@keyword
def generate_fake_name(length=8):
    vowels = 'aeiou'
    consonants = ''.join(set(string.ascii_lowercase) - set(vowels))
    fake_name = ''.join(random.choice(consonants) + random.choice(vowels) for _ in range(length // 2))
    return fake_name.capitalize()

@keyword
def generate_random_password():
    # Define the character sets
    lowercase_letters = string.ascii_lowercase
    uppercase_letters = string.ascii_uppercase
    digits = string.digits
    special_characters = string.punctuation
    # Combine all character sets
    all_characters = lowercase_letters + uppercase_letters + digits + special_characters
    # Make sure the password length is between 6 and 100 characters
    password_length = random.randint(9, 12)

    # Ensure the password has at least one of each type of character
    password = [
        random.choice(lowercase_letters),
        random.choice(uppercase_letters),
        random.choice(digits),
        random.choice(special_characters)
    ]
    # Fill the rest of the password with random characters
    for _ in range(password_length - 4):
        password.append(random.choice(all_characters))
    # Shuffle the characters to make the password random
    random.shuffle(password)

    # Convert the list of characters to a string
    password = ''.join(password)
    return password

@keyword
def remove_characters(data):
    # Define characters to remove
    chars_to_remove = ["-", ",", "₹", "+"]

    # Iterate over each character to remove
    for char in chars_to_remove:
        # Remove the character from the data
        data = data.replace(char, "")

    return data.strip()  # Remove leading and trailing whitespace

@keyword
def arrange_dates_latest_first_keyword(*dates):
    formatted_dates = []
    for date_str in dates:
        try:
            formatted_dates.append(datetime.strptime(date_str, '%d %b %Y %I:%M:%S %p'))
        except ValueError:
            pass

    sorted_dates = sorted(formatted_dates, reverse=True)
    return [date.strftime('%d %b %Y %I:%M:%S %p') for date in sorted_dates]

@keyword
def split_string_with_comma_seperator():
    pass

@keyword
def arrange_dates_oldest_first_keyword(*dates):
    formatted_dates = []
    for date_str in dates:
        try:
            formatted_dates.append(datetime.strptime(date_str, '%d %b %Y %I:%M:%S %p'))
        except ValueError:
            pass

    sorted_dates = sorted(formatted_dates)
    return [date.strftime('%d %b %Y %I:%M:%S %p') for date in sorted_dates]

@keyword
def generate_valid_password_for_primary_user():
    # Define the character sets
    lowercase_letters = string.ascii_lowercase
    uppercase_letters = string.ascii_uppercase
    digits = string.digits
    special_characters = string.punctuation
    # Combine all character sets
    all_characters = lowercase_letters + uppercase_letters + digits + special_characters
    # Make sure the password length is between 6 and 100 characters
    password_length = random.randint(6, 8)

    # Ensure the password has at least one of each type of character
    password = [
        random.choice(lowercase_letters),
        random.choice(uppercase_letters),
        random.choice(digits),
        random.choice(special_characters)
    ]
    # Fill the rest of the password with random characters
    for _ in range(password_length+2):
        password.append(random.choice(all_characters))
    # Shuffle the characters to make the password random
    random.shuffle(password)

    # Convert the list of characters to a string
    password = ''.join(password)
    return password

@keyword
def generate_invalid_password_for_primary_user():
    # Define the character sets
    lowercase_letters = string.ascii_lowercase
    uppercase_letters = string.ascii_uppercase
    digits = string.digits
    special_characters = string.punctuation
    # Combine all character sets
    wrong_combination_list = random.sample([lowercase_letters, uppercase_letters, digits, special_characters], 3)

    password_length = random.randint(6, 8)
    # password=[]
    # Ensure the password has at least one of the condition missing
    password = [random.choice(wrong_combination_list[0]), random.choice(wrong_combination_list[1]), random.choice(wrong_combination_list[2])]
    for _ in range(password_length-6):
        password.append(random.choice(wrong_combination_list))
    # Shuffle the characters to make the password random
    random.shuffle(password)

    # Convert the list of characters to a string
    p2 = ''.join(map(str, password))
    return p2